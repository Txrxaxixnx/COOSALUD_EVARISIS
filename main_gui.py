# main_gui.py:
import tkinter as tk
from tkinter import messagebox, scrolledtext, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from PIL import Image, ImageTk
import os
import sys
import subprocess
import threading
from datetime import datetime, timedelta
import requests
import time
import pandas as pd
import glob
import re
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
import sqlite3
from ttkbootstrap.dialogs import Messagebox
import json

try:
    import notion_control_interno
    import glosas_downloader
    from calendario import CalendarioInteligente
    import db_manager
except ImportError as e:
    print(f"Error de importación: {e}")
    notion_control_interno = None
    glosas_downloader = None
    CalendarioInteligente = None
    db_manager = None

def get_base_path():
    """
    Ruta base robusta para recursos empaquetados (PyInstaller onefile/onedir) y modo script.
    """
    if getattr(sys, 'frozen', False):
        # PyInstaller
        # 1) _MEIPASS (onefile), 2) exe dir/_internal (onedir personalizado), 3) exe dir
        base_dir = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
        internal = os.path.join(base_dir, '_internal')
        return internal if os.path.isdir(internal) else base_dir
    else:
        return os.path.dirname(os.path.abspath(__file__))


class CoosaludApp(ttk.Window):
    def __init__(self, nombre_usuario, cargo_usuario, foto_path, tema):
        super().__init__(themename=tema)
        self.server_process = None
        self.driver_glosas = None
        self.download_dir_glosas = None
        self.glosas_thread_lock = threading.Lock()
        self.last_processed_glosa_id = None
        self.resultados_actuales = []
        self.db_viewer_window = None
        self.current_user = {"nombre": nombre_usuario, "cargo": cargo_usuario}
        self.foto_usuario = self._cargar_foto_usuario(foto_path)
        self.title("EVARISIS GESTOR COOSALUD")
        self.state('zoomed')
        self.base_path = get_base_path()
        try:
            icon_path = os.path.join(self.base_path, "gestorcoosalud.ico")
            self.iconbitmap(icon_path)
        except Exception as e:
            print(f"Advertencia: No se pudo cargar el ícono de la aplicación: {e}")
        try:
            import configparser
            config = configparser.ConfigParser()
            config.read(os.path.join(self.base_path, 'config.ini'))            
            self.NOTION_API_KEY = config['Notion']['ApiKey']
            self.NOTION_SESSION_PAGE_ID = config['Notion']['NOTION_SESSION_PAGE_ID']            
            self.chrome_driver_path = os.path.join(self.base_path, "chrome-win64", "chromedriver.exe")
            self.chrome_binary_path = os.path.join(self.base_path, "chrome-win64", "chrome.exe")
        except (KeyError, configparser.Error) as e:
            messagebox.show_error("Error de Configuración", f"No se pudo leer 'config.ini' o falta una clave.\nDetalles: {e}")
            self.destroy()
            return        
        self.image_path = os.path.join(self.base_path, "imagenes")
        self.COLOR_AZUL_HUV = "#005A9C"
        self.logos = self._cargar_logos()
        self._configurar_estilos()
        self._crear_widgets()
        self.after(500, self.comprobar_estado_servidor)
        self.after(1000, self.ejecutar_control_interno)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _cargar_logos(self):
        logos = {}
        logos_a_cargar = {
            "huv": "logo1.jpg", "coosalud": "logo1.png",
            "innovacion_cartoon": "logo2.png", "innovacion_moderno": "logo3.png"
        }
        for nombre, archivo in logos_a_cargar.items():
            try:
                ruta_completa = os.path.join(self.image_path, archivo)
                size = (120, 40) if nombre == "coosalud" else (80, 80)
                img = Image.open(ruta_completa).resize(size, Image.Resampling.LANCZOS)
                logos[nombre] = ImageTk.PhotoImage(img)
            except Exception: logos[nombre] = None
        return logos

    def _cargar_foto_usuario(self, foto_path):
        if foto_path and foto_path != "SIN_FOTO" and os.path.exists(foto_path):
            try:
                img_header = Image.open(foto_path).resize((60, 60), Image.Resampling.LANCZOS)
                img_sidebar = Image.open(foto_path).resize((48, 48), Image.Resampling.LANCZOS)
                return {
                    "header": ImageTk.PhotoImage(img_header),
                    "sidebar": ImageTk.PhotoImage(img_sidebar)
                }
            except Exception as e:
                print(f"Error al cargar foto de usuario: {e}")
        return None
    
    def _configurar_estilos(self):
        style = self.style
        style.configure('Sidebar.TButton', font=('Segoe UI', 12), anchor='w')
        style.configure('Header.TLabel', font=('Segoe UI', 22, "bold"), foreground=self.COLOR_AZUL_HUV)
        style.configure('Status.TLabel', font=('Segoe UI', 11, 'bold'), padding=10)

    def _crear_widgets(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self._crear_cabecera()
        self._crear_menu_lateral()
        self._crear_panel_principal()
        self._crear_barra_estado()
    
    def _crear_cabecera(self):
        header_frame = ttk.Frame(self, padding=(20, 10), style='primary.TFrame')
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")        
        if self.logos.get("huv"): 
            ttk.Label(header_frame, image=self.logos["huv"], style='primary.TLabel').pack(side=LEFT, padx=(0, 15))        
        ttk.Label(header_frame, text="EVARISIS GESTOR COOSALUD", style='Header.TLabel', background=self.style.colors.primary).pack(side=LEFT, expand=True)        
        if self.foto_usuario:
            ttk.Label(header_frame, image=self.foto_usuario["header"], style='primary.TLabel').pack(side=RIGHT, padx=10)
        profile_info_frame = ttk.Frame(header_frame, style='primary.TFrame')
        profile_info_frame.pack(side=RIGHT, padx=(0, 10))
        ttk.Label(profile_info_frame, text=self.current_user["nombre"], font=('Segoe UI', 12, "bold"), anchor=E, style='primary.TLabel').pack(fill=X)
        ttk.Label(profile_info_frame, text=self.current_user["cargo"], font=('Segoe UI', 9), anchor=E, style='secondary.TLabel').pack(fill=X)

    def _crear_menu_lateral(self):
        # MODIFICADO: Reestructuramos el menú lateral para la nueva info
        sidebar_frame = ttk.Frame(self, padding=20, width=250)
        sidebar_frame.grid(row=1, column=0, sticky="ns")
        sidebar_frame.grid_propagate(False)
        
        # --- Frame de estado general ---
        frame_estado = ttk.Frame(sidebar_frame)
        frame_estado.pack(fill=X, pady=(0, 15))
        self.lbl_estado_servidor = ttk.Label(frame_estado, text=" ⚪ Comprobando...", bootstyle="secondary", font=('Segoe UI', 11, 'bold'), padding=10)
        self.lbl_estado_servidor.pack(fill=X)

        # --- NUEVO: Frame de Información del Servidor ---
        self.info_servidor_frame = ttk.Labelframe(sidebar_frame, text="Información del Servidor", padding=10)
        self.info_servidor_frame.pack(fill=X, pady=(0, 20))
        
        # Contenedor para la foto y el texto
        info_content_frame = ttk.Frame(self.info_servidor_frame)
        info_content_frame.pack(fill=X)

        # Widget para la foto
        self.lbl_foto_servidor = ttk.Label(info_content_frame)
        if self.foto_usuario:
            self.lbl_foto_servidor.config(image=self.foto_usuario["sidebar"])
        self.lbl_foto_servidor.pack(side=LEFT, padx=(0, 10))
        
        # Contenedor para los labels de texto
        info_text_frame = ttk.Frame(info_content_frame)
        info_text_frame.pack(side=LEFT, fill=X, expand=True)
        
        self.lbl_usuario_servidor = ttk.Label(info_text_frame, text="Iniciado por: N/A", font=('Segoe UI', 9, 'bold'))
        self.lbl_usuario_servidor.pack(fill=X, anchor='w')
        self.lbl_update_servidor = ttk.Label(info_text_frame, text="Últ. Act: --:--:--", font=('Segoe UI', 8))
        self.lbl_update_servidor.pack(fill=X, anchor='w')

        # Ocultamos el frame al inicio
        self.info_servidor_frame.pack_forget()

        # --- Botones de acción ---
        self.btn_iniciar_sesion_cliente = ttk.Button(sidebar_frame, text="  Iniciar Sesión Cliente", style='Sidebar.TButton', command=self.iniciar_sesion_cliente)
        self.btn_iniciar_sesion_cliente.pack(fill=X, pady=5, ipady=10)
        
        self.btn_iniciar_servidor = ttk.Button(sidebar_frame, text="  Iniciar Servidor", style='Sidebar.TButton', command=self.iniciar_servidor, state="disabled")
        self.btn_iniciar_servidor.pack(fill=X, pady=5, ipady=10)
        
        # --- Botones de navegación ---
        ttk.Separator(sidebar_frame, orient=HORIZONTAL).pack(fill=X, pady=15)
        ttk.Button(sidebar_frame, text="  Dashboard", style='Sidebar.TButton', command=lambda: self.mostrar_panel("dashboard")).pack(fill=X, pady=5, ipady=10)
        ttk.Button(sidebar_frame, text="  Configuración", style='Sidebar.TButton', command=lambda: self.mostrar_panel("configuracion")).pack(fill=X, pady=5, ipady=10)
        ttk.Button(sidebar_frame, text="  Bases de datos", style='Sidebar.TButton', command=lambda: self.mostrar_panel("bases_de_datos")).pack(fill=X, pady=5, ipady=10)
       
    def _crear_panel_principal(self):
        self.main_panel = ttk.Frame(self, padding=(20,20,20,0))
        self.main_panel.grid(row=1, column=1, sticky="nsew")
        self.main_panel.grid_rowconfigure(0, weight=1)
        self.main_panel.grid_columnconfigure(0, weight=1)

        self.paneles = {}
        self.paneles["bienvenida"] = self._crear_panel_bienvenida()
        self.paneles["dashboard"] = self._crear_panel_dashboard()
        self.paneles["configuracion"] = self._crear_panel_configuracion()
        self.paneles["bases_de_datos"] = self._crear_panel_bases_de_datos()

        self.mostrar_panel("bienvenida")

    def _accion_bases_de_datos(self):
        """Muestra un mensaje temporal para la sección de Bases de Datos."""
        Messagebox.show_info(
            title="En Construcción",
            message="Esta sección se encuentra actualmente en desarrollo.",
            parent=self
        )

    def _crear_panel_bienvenida(self):
        panel = ttk.Frame(self.main_panel)
        panel.grid(row=0, column=0, sticky="nsew")
        content_frame = ttk.Frame(panel)
        content_frame.place(relx=0.5, rely=0.5, anchor="center")
        ttk.Label(content_frame, text="Bienvenido/a", font=("Segoe UI", 20, "bold")).pack(pady=10)
        info_text = "La herramienta está iniciando los controles internos.\nPor favor, espere mientras se verifica el estado de la sesión..."
        ttk.Label(content_frame, text=info_text, font=("Segoe UI", 12), justify="center").pack(pady=10, fill=X)
        return panel

    def _crear_panel_configuracion(self):
        panel = ttk.Frame(self.main_panel, padding=(0, 0, 0, 10))
        panel.grid(row=0, column=0, sticky="nsew")
        panel.grid_rowconfigure(1, weight=1)
        panel.grid_columnconfigure(0, weight=1)

        content_frame = ttk.Frame(panel)
        content_frame.grid(row=0, column=0, sticky="new", pady=(0, 20))
        ttk.Label(content_frame, text="Panel de Configuración", font=("Segoe UI", 20, "bold")).pack(pady=10)
        ttk.Label(content_frame, text="Aquí se mostrará el registro de procesos y futuras opciones.", font=("Segoe UI", 12)).pack()
        
        console_frame = ttk.Labelframe(panel, text="Registro de Proceso", padding=10)
        console_frame.grid(row=1, column=0, sticky="nsew")

        self.console_text = scrolledtext.ScrolledText(console_frame, height=8, state="disabled", wrap=tk.WORD, font=("Consolas", 10))
        self.console_text.pack(fill="both", expand=True)
        
        return panel

    def _crear_panel_bases_de_datos(self):
        panel = ttk.Frame(self.main_panel)
        panel.grid_rowconfigure(2, weight=1)
        panel.grid_columnconfigure(0, weight=1)

        controles_frame = ttk.Frame(panel)
        controles_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(controles_frame, text="Fecha Inicial:", font="-size 10").pack(side=LEFT, padx=(0, 5))
        self.db_fecha_ini_var = tk.StringVar(value="No seleccionada")
        ttk.Label(controles_frame, textvariable=self.db_fecha_ini_var, bootstyle="info", padding=5).pack(side=LEFT, padx=(0, 5))
        ttk.Button(controles_frame, text="Seleccionar...", command=self._seleccionar_fecha_db_ini, bootstyle="secondary-outline").pack(side=LEFT, padx=(0, 20))

        ttk.Label(controles_frame, text="Fecha Final:", font="-size 10").pack(side=LEFT, padx=(0, 5))
        self.db_fecha_fin_var = tk.StringVar(value="No seleccionada")
        ttk.Label(controles_frame, textvariable=self.db_fecha_fin_var, bootstyle="info", padding=5).pack(side=LEFT, padx=(0, 5))
        ttk.Button(controles_frame, text="Seleccionar...", command=self._seleccionar_fecha_db_fin, bootstyle="secondary-outline").pack(side=LEFT, padx=(0, 20))

        self.btn_db_buscar = ttk.Button(controles_frame, text="Buscar y Cargar en BD", command=self.iniciar_proceso_db, state="disabled")
        self.btn_db_buscar.pack(side=LEFT, padx=10, ipady=4)

        # <<<----------- NUEVO BOTÓN AÑADIDO Y CONFIGURADO AQUÍ -----------<<<
        self.btn_db_visualizar = ttk.Button(controles_frame, text="Visualizar Base de Datos", command=self._visualizar_base_de_datos, state="disabled")
        self.btn_db_visualizar.pack(side=LEFT, padx=10, ipady=4)
        
        self.db_progress_frame = ttk.Frame(panel, padding=(0, 5))
        self.db_progress_frame.grid(row=1, column=0, sticky="ew")
        
        self.db_progress_label = ttk.Label(self.db_progress_frame, text="Cargando 0/0...")
        self.db_progress_label.pack(side=LEFT, padx=(0, 10))
        
        self.db_progressbar = ttk.Progressbar(self.db_progress_frame, mode='determinate')
        self.db_progressbar.pack(side=LEFT, fill=X, expand=True)

        self.db_results_frame = ttk.Frame(panel)
        self.db_results_frame.grid(row=2, column=0, sticky="nsew")

        return panel
    
    # <<<----------- NUEVA FUNCIÓN AÑADIDA AQUÍ -----------<<<
    def _visualizar_base_de_datos(self):
        """Abre una ventana con el contenido de glosas_coosalud.db."""
        if not db_manager:
            Messagebox.show_error("Error de Módulo", "El módulo 'db_manager.py' no se pudo cargar.", parent=self)
            return

        db_path = os.path.join(self.base_path, db_manager.DB_FILENAME)
        if not os.path.exists(db_path):
            Messagebox.show_warning(
                "Base de datos no encontrada",
                f"No se encontró el archivo {db_manager.DB_FILENAME}.",
            )
            return

        if self.db_viewer_window and self.db_viewer_window.winfo_exists():
            self.db_viewer_window.focus()
            self.db_viewer_window.lift()
            return

        conn = None
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cuentas_rows = conn.execute(
                "SELECT id_cuenta, radicacion, fecha_rad, factura, valor_factura, valor_glosado "
                "FROM cuentas ORDER BY fecha_rad DESC, factura DESC"
            ).fetchall()
            glosas_rows = conn.execute(
                "SELECT factura, id_item, descripcion_item, tipo, descripcion, valor_glosado_item, "
                "usuario, fecha_glosa, estado FROM glosas_detalle "
                "ORDER BY fecha_glosa DESC, factura DESC"
            ).fetchall()
            cuentas = [dict(row) for row in cuentas_rows]
            glosas = [dict(row) for row in glosas_rows]
        except sqlite3.Error as e:
            Messagebox.show_error(
                f"No se pudo consultar la base de datos.\n\n{e}",
                "Error al leer la base de datos",
                parent=self
            )
            return
        finally:
            if conn:
                conn.close()

        self.db_viewer_window = tk.Toplevel(self)
        self.db_viewer_window.title("Registros guardados en glosas_coosalud.db")
        self.db_viewer_window.geometry("1020x640")
        self.db_viewer_window.transient(self)

        def _on_close():
            if self.db_viewer_window:
                self.db_viewer_window.destroy()
                self.db_viewer_window = None

        self.db_viewer_window.protocol("WM_DELETE_WINDOW", _on_close)

        def _handle_destroy(event):
            if self.db_viewer_window and event.widget == self.db_viewer_window:
                self.db_viewer_window = None

        self.db_viewer_window.bind("<Destroy>", _handle_destroy)

        container = ttk.Frame(self.db_viewer_window, padding=15)
        container.pack(fill=tk.BOTH, expand=True)

        notebook = ttk.Notebook(container)
        notebook.pack(fill=tk.BOTH, expand=True)

        width_map = {
            "id_cuenta": 150,
            "radicacion": 140,
            "fecha_rad": 110,
            "factura": 120,
            "valor_factura": 140,
            "valor_glosado": 140,
            "descripcion_item": 260,
            "tipo": 120,
            "descripcion": 260,
            "valor_glosado_item": 150,
            "usuario": 120,
            "fecha_glosa": 150,
            "estado": 120,
        }

        max_rows = 800

        def _crear_tab(titulo, columnas, datos):
            marco = ttk.Frame(notebook, padding=10)
            notebook.add(marco, text=titulo)
            marco.grid_rowconfigure(1, weight=1)
            marco.grid_columnconfigure(0, weight=1)

            total = len(datos)
            datos_mostrados = datos[:max_rows]
            ttk.Label(
                marco,
                text=f"Mostrando {len(datos_mostrados)} de {total} registros",
                anchor="w"
            ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

            tree = ttk.Treeview(marco, columns=columnas, show="headings")
            vsb = ttk.Scrollbar(marco, orient=tk.VERTICAL, command=tree.yview)
            hsb = ttk.Scrollbar(marco, orient=tk.HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            tree.grid(row=1, column=0, sticky="nsew")
            vsb.grid(row=1, column=1, sticky="ns")
            hsb.grid(row=2, column=0, sticky="ew")

            for col in columnas:
                cabecera = col.replace("_", " ").title()
                tree.heading(col, text=cabecera)
                tree.column(col, width=width_map.get(col, 120), anchor=tk.W)

            for fila in datos_mostrados:
                valores = ["" if fila.get(col) is None else str(fila.get(col)) for col in columnas]
                tree.insert("", tk.END, values=valores)

        _crear_tab(
            "Cuentas",
            ["id_cuenta", "radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"],
            cuentas
        )

        _crear_tab(
            "Detalle de glosas",
            [
                "factura",
                "id_item",
                "descripcion_item",
                "tipo",
                "descripcion",
                "valor_glosado_item",
                "usuario",
                "fecha_glosa",
                "estado"
            ],
            glosas
        )

        notebook.enable_traversal()
        self.db_viewer_window.focus()

    def _seleccionar_fecha_db_ini(self):
        fecha = CalendarioInteligente.seleccionar_fecha(parent=self, titulo="Seleccione la FECHA DE INICIO")
        if fecha:
            self.fecha_ini_db = fecha
            self.db_fecha_ini_var.set(fecha.strftime("%d/%m/%Y"))
            self._validar_fechas_db()

    def _seleccionar_fecha_db_fin(self):
        fecha_inicial = getattr(self, 'fecha_ini_db', None)
        fecha = CalendarioInteligente.seleccionar_fecha(parent=self, titulo="Seleccione la FECHA DE FIN", fecha_inicial=fecha_inicial)
        if fecha:
            self.fecha_fin_db = fecha
            self.db_fecha_fin_var.set(fecha.strftime("%d/%m/%Y"))
            self._validar_fechas_db()
            
    def _validar_fechas_db(self):
        if hasattr(self, 'fecha_ini_db') and hasattr(self, 'fecha_fin_db'):
            if self.fecha_fin_db < self.fecha_ini_db:
                Messagebox.show_error("La fecha de fin no puede ser anterior a la fecha de inicio.", "Error de Fechas", parent=self)
                self.btn_db_buscar.config(state="disabled")
            else:
                self.btn_db_buscar.config(state="normal")
        else:
            self.btn_db_buscar.config(state="disabled")

    def iniciar_proceso_db(self):
        if not db_manager:
            Messagebox.show_error("Error de Módulo", "El módulo 'db_manager.py' no se pudo cargar.", parent=self)
            return

        fecha_ini_str = self.fecha_ini_db.strftime("%Y-%m-%d")
        fecha_fin_str = self.fecha_fin_db.strftime("%Y-%m-%d")

        self._log_to_console(f"\n[BD] Iniciando búsqueda para base de datos entre {fecha_ini_str} y {fecha_fin_str}...")
        
        threading.Thread(
            target=self._tarea_buscar_para_db, 
            args=(fecha_ini_str, fecha_fin_str), 
            daemon=True
        ).start()


# En main_gui.py, reemplace la función completa _tarea_buscar_para_db

    def _tarea_buscar_para_db(self, fecha_ini, fecha_fin):
        if not self.glosas_thread_lock.acquire(blocking=False):
            self.after(0, lambda: Messagebox.show_warning("Ya hay otra tarea en ejecución. Por favor, espere.", "Tarea en Progreso"))
            return

        try:
            self.after(0, lambda: self.btn_db_buscar.config(state="disabled"))
            self.after(0, lambda: self.btn_db_visualizar.config(state="disabled"))

            if not self._iniciar_navegador_glosas_si_no_existe():
                self.glosas_thread_lock.release()
                return
            
            self.after(0, self._preparar_ui_para_carga_db)
            
            from glosas_downloader import establecer_contexto_busqueda, extraer_datos_tabla_actual, cambiar_numero_entradas
            
            def progress_callback(actual, total, data_fila):
                self.after(0, self._actualizar_progreso_db, actual, total, data_fila)

            self.after(0, lambda: self._log_to_console("[BD] Obteniendo la lista completa de facturas del portal..."))
            establecer_contexto_busqueda(self.driver_glosas, fecha_ini, fecha_fin)
            cambiar_numero_entradas(self.driver_glosas, "-1")
            
            cuentas_a_procesar, _ = extraer_datos_tabla_actual(self.driver_glosas, progress_callback)
            
            total_encontradas = len(cuentas_a_procesar)
            self.after(0, lambda: self._log_to_console(f"[BD] Mapeo finalizado. Se encontraron {total_encontradas} facturas."))

            if not cuentas_a_procesar:
                self.after(0, lambda: self._log_to_console("[BD] No hay facturas para procesar."))
                self.after(0, self._finalizar_carga_db)
                # No liberamos el lock aquí, se libera en el finally
                return

            # <<<----------- CORRECCIÓN IMPORTANTE AQUÍ -----------<<<
            # Movemos la lógica de la pregunta a una función separada que se ejecuta después
            # de que el mapeo haya terminado, para no bloquear el hilo de la UI.
            self.after(100, self._preguntar_y_lanzar_procesamiento_db, cuentas_a_procesar, fecha_ini, fecha_fin)

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [BD] Error en la tarea principal de BD: {e}")
            self.after(0, self._finalizar_carga_db)
            if self.glosas_thread_lock.locked():
                self.glosas_thread_lock.release()
        # El finally se quita de aquí y se pone en la nueva función que realmente termina el proceso.
        
    # <<<----------- NUEVA FUNCIÓN PARA GESTIONAR EL FLUJO CORRECTAMENTE -----------<<<
    def _preguntar_y_lanzar_procesamiento_db(self, cuentas_a_procesar, fecha_ini, fecha_fin):
        """
        Esta función se ejecuta en el hilo principal de la UI, pregunta al usuario
        y luego lanza el proceso pesado en un nuevo hilo.
        """
        total_encontradas = len(cuentas_a_procesar)

        cantidad_str = simpledialog.askstring(
            "Confirmar Cantidad a Procesar",
            f"Se encontraron {total_encontradas} facturas en total.\n\n"
            f"¿Cuántas desea descargar y guardar en la base de datos?\n"
            f"(Escriba un número o la palabra 'todas')",
            parent=self
        )

        if not cantidad_str:
            self._log_to_console("[BD] Proceso cancelado por el usuario.")
            self._finalizar_carga_db()
            self.glosas_thread_lock.release()
            return

        try:
            cantidad_str = cantidad_str.strip().lower()
            if cantidad_str == 'todas':
                cantidad_a_procesar = total_encontradas
            else:
                cantidad_a_procesar = int(cantidad_str)
                if not (0 < cantidad_a_procesar <= total_encontradas):
                    Messagebox.show_error("Cantidad Inválida", f"Por favor, ingrese un número entre 1 y {total_encontradas}.", parent=self)
                    self._finalizar_carga_db()
                    self.glosas_thread_lock.release()
                    return
        except ValueError:
            Messagebox.show_error("Entrada Inválida", "Por favor, ingrese un número válido o la palabra 'todas'.", parent=self)
            self._finalizar_carga_db()
            self.glosas_thread_lock.release()
            return

        cuentas_finales = cuentas_a_procesar[:cantidad_a_procesar]
        
        # Ahora que tenemos la cantidad, lanzamos el proceso de descarga en un nuevo hilo
        threading.Thread(
            target=self._tarea_procesamiento_lote_db,
            args=(cuentas_finales, fecha_ini, fecha_fin),
            daemon=True
        ).start()

    def _tarea_procesamiento_lote_db(self, cuentas_finales, fecha_ini, fecha_fin):
        """Esta es la tarea pesada que corre en su propio hilo."""
        try:
            self.after(0, lambda: self._log_to_console(f"\n[BD] Confirmado. Se procesarán {len(cuentas_finales)} facturas."))
            self.after(0, lambda: self._log_to_console("[BD] Iniciando descarga y guardado en base de datos..."))
            
            def log_desde_hilo(mensaje):
                self.after(0, self._log_to_console, mensaje)

            proceso_exitoso = db_manager.procesar_cuentas_en_lote(
                self.driver_glosas, cuentas_finales, self.base_path, 
                log_desde_hilo, fecha_ini, fecha_fin, self.download_dir_glosas
            )

            if proceso_exitoso:
                self.after(0, lambda: self.btn_db_visualizar.config(state="normal"))
                self.after(0, lambda: Messagebox.show_info("Proceso Completado", f"Se han procesado y guardado {len(cuentas_finales)} facturas.", parent=self))
            else:
                self.after(0, lambda: Messagebox.show_error("Proceso Interrumpido", "Error durante el procesamiento.", parent=self))
            
            self.after(0, self._finalizar_carga_db)
        
        except Exception as e:
            self.after(0, self._log_to_console(f"❌ Error en la tarea de procesamiento en lote: {e}"))
            self.after(0, self._finalizar_carga_db)
        finally:
            if self.glosas_thread_lock.locked():
                self.glosas_thread_lock.release()
            self.after(0, self._log_to_console, "[BD] Bloqueo liberado.")


    # <<<----------- Y AÑADIMOS ESTAS TRES NUEVAS FUNCIONES DE AYUDA -----------<<<
    def _preparar_ui_para_carga_db(self):
        """Limpia la tabla y muestra la barra de progreso."""
        for widget in self.db_results_frame.winfo_children():
            widget.destroy()
            
        self.db_progress_frame.grid(row=1, column=0, sticky="ew")
        
        cols = ["radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"]
        self.db_tree = ttk.Treeview(self.db_results_frame, columns=cols, show='headings')
        self.db_tree.heading("radicacion", text="N° Radicación")
        self.db_tree.heading("fecha_rad", text="Fecha Rad.")
        self.db_tree.heading("factura", text="N° Factura")
        self.db_tree.heading("valor_factura", text="Valor Factura")
        self.db_tree.heading("valor_glosado", text="Valor Glosado")
        self.db_tree.pack(fill="both", expand=True)

    def _actualizar_progreso_db(self, actual, total, data_fila):
        """Actualiza la barra, el contador y añade una fila a la tabla."""
        if total > 0:
            self.db_progressbar['maximum'] = total
            self.db_progressbar['value'] = actual
            self.db_progress_label.config(text=f"Mapeando {actual}/{total}...")
        
        # Si nos pasaron datos de una fila, la añadimos a la tabla en tiempo real
        if data_fila:
            self.db_tree.insert("", "end", values=(
                data_fila['radicacion'], data_fila['fecha_rad'], data_fila['factura'],
                data_fila['valor_factura'], data_fila['valor_glosado']
            ))

    def _finalizar_carga_db(self):
        """Oculta la barra de progreso y reactiva el botón de búsqueda."""
        self.db_progress_frame.grid_forget()
        self.btn_db_buscar.config(state="normal")


    def _mostrar_resultados_db(self, resultados):
        # Limpiamos el frame de resultados por si había una búsqueda anterior
        for widget in self.db_results_frame.winfo_children():
            widget.destroy()
        
        if not resultados:
            ttk.Label(self.db_results_frame, text="No se encontraron resultados para el rango seleccionado.").pack(pady=20)
            return

        # Creamos la tabla (Treeview) para mostrar los resultados
        cols = ["radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"]
        tree = ttk.Treeview(self.db_results_frame, columns=cols, show='headings')
        
        tree.heading("radicacion", text="N° Radicación")
        tree.heading("fecha_rad", text="Fecha Rad.")
        tree.heading("factura", text="N° Factura")
        tree.heading("valor_factura", text="Valor Factura")
        tree.heading("valor_glosado", text="Valor Glosado")
        
        for res in resultados:
            tree.insert("", "end", values=(res['radicacion'], res['fecha_rad'], res['factura'], res['valor_factura'], res['valor_glosado']))
        
        tree.pack(fill="both", expand=True)

    def _crear_panel_dashboard(self):
        panel = ttk.Frame(self.main_panel)
        panel.grid(row=0, column=0, sticky="nsew")
        panel.grid_rowconfigure(3, weight=1) # Fila para la tabla de resultados
        panel.grid_columnconfigure(0, weight=1)

        ttk.Label(panel, text="Dashboard de Tareas Automatizadas", font=("Segoe UI", 20, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky='w')
        
        self.btn_glosas = ttk.Button(panel, text="Buscar Informes de Glosas", command=self.iniciar_proceso_glosas, bootstyle="success", state="disabled")
        self.btn_glosas.grid(row=1, column=0, columnspan=2, sticky="ew", pady=5)

        # ## CAMBIO: Se eliminó el botón de automatización completa de aquí ##
        # self.btn_automatizar_glosas = ttk.Button(...) # <-- LÍNEA ELIMINADA

        self.results_frame = ttk.Frame(panel, padding=(0, 20, 0, 0))
        self.results_frame.grid(row=3, column=0, columnspan=2, sticky="nsew")
        
        # ## SECCIÓN DE PAGINACIÓN Y AUTOMATIZACIÓN ##
        self.pagination_frame = ttk.Frame(panel, padding=(0, 10, 0, 0))
        self.pagination_frame.grid(row=4, column=0, columnspan=2, sticky="ew")
        self.pagination_frame.grid_columnconfigure(3, weight=1) # Columna del medio se expande

        # Controles de paginación
        self.combo_entradas = ttk.Combobox(self.pagination_frame, values=["20", "50", "100", "500", "Todos"], state="readonly", width=8)
        self.combo_entradas.bind("<<ComboboxSelected>>", self.on_cambiar_entradas)
        self.combo_entradas.pack(side=LEFT, padx=5)

        self.btn_anterior = ttk.Button(self.pagination_frame, text="Anterior", command=lambda: self.on_navegar("anterior"))
        self.btn_anterior.pack(side=LEFT, padx=5)
        
        self.lbl_paginacion_info = ttk.Label(self.pagination_frame, text="", anchor="center")
        self.lbl_paginacion_info.pack(side=LEFT, padx=10, expand=True, fill=X)

        # ## NUEVO: Botón de automatización integrado en la barra de resultados ##
        self.btn_iniciar_automatizacion = ttk.Button(
            self.pagination_frame, 
            text="🤖 Iniciar Automatización", 
            command=self.iniciar_proceso_automatizacion_integrada, # Llama a la nueva función
            bootstyle="warning"
        )
        self.btn_iniciar_automatizacion.pack(side=RIGHT, padx=5)
        
        self.btn_siguiente = ttk.Button(self.pagination_frame, text="Siguiente", command=lambda: self.on_navegar("siguiente"))
        self.btn_siguiente.pack(side=RIGHT, padx=5)

        self.pagination_frame.grid_remove() # Ocultamos todo el panel al inicio
        
        return panel

    def mostrar_panel(self, nombre_panel):
        for panel in self.paneles.values():
            panel.grid_forget()
        self.paneles[nombre_panel].grid(row=0, column=0, sticky="nsew")

    def _crear_barra_estado(self):
        status_bar = ttk.Frame(self, padding=(10, 5), style='secondary.TFrame')
        status_bar.grid(row=3, column=0, columnspan=2, sticky="ew")
        ttk.Label(status_bar, text="Desarrollado por Innovación y Desarrollo del HUV Versión 2.0", font=("Segoe UI", 8), style='secondary.TLabel').pack(side=LEFT)
        self.lbl_reloj = ttk.Label(status_bar, font=("Segoe UI", 8), style='secondary.TLabel')
        self.lbl_reloj.pack(side=RIGHT)
        self._actualizar_reloj()

    def _actualizar_reloj(self):
        now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        self.lbl_reloj.config(text=now)
        self.after(1000, self._actualizar_reloj)

    def _log_to_console(self, message):
        self.console_text.config(state="normal")
        self.console_text.insert(tk.END, message + "\n")
        self.console_text.see(tk.END)
        self.console_text.config(state="disabled")

    def on_closing(self):
        self._log_to_console("Cerrando aplicación...")
        if self.server_process and self.server_process.poll() is None:
            self._log_to_console("[GUI] Cerrando el proceso del servidor en segundo plano...")
            self.server_process.terminate()
            self.server_process.wait()
            self._log_to_console("[GUI] Proceso del servidor cerrado.")
        
        if self.driver_glosas:
            self._log_to_console("[Glosas] Cerrando navegador de automatización...")
            try:
                self.driver_glosas.quit()
            except Exception as e:
                self._log_to_console(f"Error al cerrar el driver de glosas: {e}")
        self.destroy()

    # ... (Las funciones de comprobación de estado, control interno, iniciar servidor/cliente no cambian)
    def comprobar_estado_servidor(self):
        threading.Thread(target=self._tarea_comprobar_estado, daemon=True).start()

    def _tarea_comprobar_estado(self):
        headers = { "Authorization": f"Bearer {self.NOTION_API_KEY}", "Notion-Version": "2022-06-28" }
        url = f"https://api.notion.com/v1/blocks/{self.NOTION_SESSION_PAGE_ID}/children"
        try:
            res = requests.get(url, headers=headers, timeout=10)
            res.raise_for_status()
            data = res.json()
            texto_bloque = None
            for block in data.get("results", []):
                if block.get("type") == "paragraph":
                    rt = block.get("paragraph", {}).get("rich_text", [])
                    if rt and rt[0].get("plain_text", "").startswith("Session PHPSESSID:"):
                        texto_bloque = rt[0].get("plain_text")
                        break
            if not texto_bloque:
                self.after(0, self._actualizar_ui_estado, "danger", "✖ Sesión Inactiva", True)
                return
            
            parts = [p.strip() for p in texto_bloque.split('|')]
            timestamp_str = parts[1].replace("LastUpdate:", "").strip()
            username = parts[2].replace("User:", "").strip() if len(parts) > 2 else "Desconocido"

            last_update = datetime.fromisoformat(timestamp_str)

            if datetime.now() - last_update > timedelta(minutes=5):
                bootstyle, texto, habilitar_boton = "warning", "⚠ Sesión Expirada", True
                # Si expiró, no mostramos la info del usuario
                self.after(0, self._actualizar_ui_estado, bootstyle, texto, habilitar_boton, None, None)
            else:
                bootstyle, texto, habilitar_boton = "success", "✔ Sesión Activa", False
                self.after(0, self._actualizar_ui_estado, bootstyle, texto, habilitar_boton, username, last_update)

        except requests.RequestException:
            self.after(0, self._actualizar_ui_estado, "danger", "✖ Error de Red", True, None, None)
        except (IndexError, ValueError) as e:
            self.after(0, self._log_to_console, f"Error parseando bloque de Notion: {e}")
            self.after(0, self._actualizar_ui_estado, "danger", "✖ Bloque Malformado", True, None, None)
        finally:
            self.after(60000, self.comprobar_estado_servidor)
            
    def _actualizar_ui_estado(self, bootstyle, texto, habilitar_boton_servidor, username, update_time):
        # MODIFICADO: Gestionamos la visibilidad del nuevo panel de información
        self.lbl_estado_servidor.configure(text=texto, bootstyle=bootstyle)
        self.btn_iniciar_servidor.config(state="normal" if habilitar_boton_servidor else "disabled")
        
        if not habilitar_boton_servidor and username and update_time:
            # Estado ACTIVO: mostramos toda la info
            self.btn_iniciar_sesion_cliente.config(state="normal")
            self.btn_glosas.config(state="normal")
            
            self.info_servidor_frame.pack(fill=X, pady=(0, 20)) # Hacemos visible el frame
            self.lbl_usuario_servidor.config(text=f"Iniciado por: {username}")
            self.lbl_update_servidor.config(text=f"Últ. Act: {update_time.strftime('%H:%M:%S')}")
            
            self.lbl_estado_servidor.config(cursor="hand2")
            self.lbl_estado_servidor.bind("<Button-1>", lambda e: self.iniciar_sesion_cliente())
            self.lbl_estado_servidor.configure(text="✔ Sesión Activa (Clic para usar)")
        else:
            # Estado INACTIVO/EXPIRADO: ocultamos la info detallada
            self.btn_iniciar_sesion_cliente.config(state="disabled")
            self.btn_glosas.config(state="disabled")
            self.info_servidor_frame.pack_forget() # Ocultamos el frame
            
            self.lbl_estado_servidor.config(cursor="")
            self.lbl_estado_servidor.unbind("<Button-1>")

    def ejecutar_control_interno(self):
        if not notion_control_interno:
            self._log_to_console("❌ ERROR: El módulo 'notion_control_interno.py' no se encontró.")
            messagebox.show_error("Error Crítico", "No se encontró el módulo de control interno.")
            return
        self._log_to_console("[1/2] Iniciando control interno...")
        threading.Thread(target=self._tarea_control_interno, daemon=True).start()

    def _tarea_control_interno(self):
        exito = notion_control_interno.registrar_uso(lambda msg: self.after(0, self._log_to_console, msg),
                                                     self.base_path)
        self.after(0, self._finalizar_control_interno, exito)

    def _finalizar_control_interno(self, exito):
        if exito:
            self._log_to_console("✅ Control interno concretado.")
        else:
            self._log_to_console("❌ La fase de control interno falló.")
            messagebox.show_error("Error de Control", "No se pudo registrar el uso de la herramienta.")

    def iniciar_servidor(self):
        self.btn_iniciar_servidor.config(state="disabled")
        self._log_to_console("\n[Anfitrión] Iniciando secuencia del servidor...")
        threading.Thread(target=self._tarea_lanzar_servidor, daemon=True).start()

    def _tarea_lanzar_servidor(self):
        try:
            python_exe = sys.executable
            env = os.environ.copy()
            env["PYTHONUTF8"] = "1"

            # Obtenemos el nombre de usuario de la instancia de la GUI
            usuario_actual = self.current_user.get("nombre", "UsuarioDesconocido")
            
            command = [python_exe]
            if not getattr(sys, 'frozen', False):
                command.append(sys.argv[0])
            
            command.extend(["--run-server", f"--base-path={self.base_path}", f"--usuario={usuario_actual}"])
            
            flag_file_path = os.path.join(self.base_path, '.sync_success.flag')
            if os.path.exists(flag_file_path):
                os.remove(flag_file_path)
            self.after(0, self._log_to_console, "[Anfitrión] Lanzando el gestor de sesión del servidor...")
            self.server_process = subprocess.Popen(
                command,
                env=env,
                creationflags=subprocess.CREATE_NO_WINDOW 
            )            
            self.after(0, self._log_to_console, "[Anfitrión] ...esperando señal de sincronización exitosa del servidor...")
            timeout = 45 
            start_time = time.time()
            sync_success = False
            while time.time() - start_time < timeout:
                if os.path.exists(flag_file_path):
                    self.after(0, self._log_to_console, "✅ [Anfitrión] ¡Señal recibida! La sincronización inicial fue un éxito.")
                    os.remove(flag_file_path)
                    sync_success = True
                    break
                time.sleep(0.5)

            if sync_success:
                self.after(0, self._log_to_console, "[Anfitrión] Forzando actualización de estado de la UI...")
                self.comprobar_estado_servidor()
            else:
                self.after(0, self._log_to_console, "❌ [Anfitrión] Error: El servidor no envió la señal de éxito a tiempo.")
                self.after(0, lambda: self.btn_iniciar_servidor.config(state="normal"))
        except Exception as e:
            error_msg = f"❌ [Anfitrión] Error inesperado al lanzar la secuencia: {e}"
            self.after(0, self._log_to_console, error_msg)
            self.after(0, lambda: self.btn_iniciar_servidor.config(state="normal"))

    def iniciar_sesion_cliente(self):
        self._log_to_console("\n[Cliente] Solicitando inicio de sesión...")
        threading.Thread(target=self._tarea_iniciar_sesion_cliente, daemon=True).start()

    def _tarea_iniciar_sesion_cliente(self):
        try:
            python_exe = sys.executable
            command = [python_exe]
            if not getattr(sys, 'frozen', False):
                command.append(sys.argv[0])
            
            command.extend(["--run-client", f"--base-path={self.base_path}"])

            subprocess.Popen(
                command, 
                creationflags=subprocess.CREATE_NO_WINDOW  
            )
            self.after(0, self._log_to_console, "✅ [Cliente] Orden de inicio de sesión enviada.")
        except Exception as e:
            error_msg = f"❌ [Cliente] Error al lanzar el proceso del cliente: {e}"
            self.after(0, self._log_to_console, error_msg)
    
    # ... (El resto de las funciones de la GUI y la lógica de Glosas se definen a continuación) ...
    # ... (Las funciones de búsqueda, descarga, paginación, etc., permanecen iguales pero ahora se añade la automatización)

    def _iniciar_navegador_glosas_si_no_existe(self):
        if self.driver_glosas:
            self.after(0, self._log_to_console, "[Glosas] Reutilizando navegador ya existente.")
            return True

        try:
            self.after(0, self._log_to_console, "[Glosas] Iniciando nuevo navegador de automatización...")
            cookie = glosas_downloader.get_session_cookie(self.base_path)
            
            self.driver_glosas, self.download_dir_glosas = glosas_downloader.setup_driver(self.base_path, for_download=True)
            
            self.after(0, self._log_to_console, "[Glosas] Inyectando cookie de sesión...")
            self.driver_glosas.get("https://vco.ctamedicas.com")
            self.driver_glosas.add_cookie({'name': 'PHPSESSID', 'value': cookie})
            self.driver_glosas.refresh()
            self.after(0, self._log_to_console, "✅ [Glosas] Navegador listo y sesión iniciada.")
            return True
        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Glosas] Error fatal al iniciar el navegador: {e}")
            self.after(0, lambda: Messagebox.show_error(f"No se pudo iniciar el navegador de automatización:\n\n{e}", "Error Crítico"))
            self.driver_glosas = None
            return False

    def iniciar_proceso_glosas(self):
        if not CalendarioInteligente:
            messagebox.show_error("Error de Módulo", "El módulo 'calendario.py' no se pudo cargar.")
            return

        self._log_to_console("\n[Glosas] Solicitando fecha de inicio...")
        fecha_ini_obj = CalendarioInteligente.seleccionar_fecha(parent=self, titulo="Seleccione la FECHA DE INICIO", codigo_pais_festivos='CO', locale='es_CO')
        
        if not fecha_ini_obj:
            self._log_to_console("[Glosas] Proceso cancelado por el usuario (fecha de inicio).")
            return
        
        fecha_ini_str = fecha_ini_obj.strftime("%Y-%m-%d")
        self._log_to_console(f"[Glosas] Fecha de inicio seleccionada: {fecha_ini_str}")

        self._log_to_console("[Glosas] Solicitando fecha de fin...")
        fecha_fin_obj = CalendarioInteligente.seleccionar_fecha(parent=self, titulo="Seleccione la FECHA DE FIN", fecha_inicial=fecha_ini_obj, codigo_pais_festivos='CO', locale='es_CO')
        
        if not fecha_fin_obj:
            self._log_to_console("[Glosas] Proceso cancelado por el usuario (fecha de fin).")
            return

        if fecha_fin_obj < fecha_ini_obj:
            self._log_to_console("[Glosas] Error: La fecha de fin es anterior a la de inicio.")
            Messagebox.show_error("La fecha de fin no puede ser anterior a la fecha de inicio.", "Error de Fechas")
            return
            
        fecha_fin_str = fecha_fin_obj.strftime("%Y-%m-%d")
        self.fecha_ini_actual = fecha_ini_str
        self.fecha_fin_actual = fecha_fin_str

        if not self.glosas_thread_lock.acquire(blocking=False):
            self._log_to_console("⚠️ [Glosas] Ya hay una tarea en ejecución. Por favor, espere.")
            Messagebox.show_warning("Ya hay una tarea de glosas en ejecución. Por favor, espere a que termine.", "Tarea en Progreso")
            return

        self._log_to_console(f"[Glosas] Buscando registros entre {fecha_ini_str} y {fecha_fin_str}...")
        hilo_busqueda = threading.Thread(target=self._tarea_buscar_glosas, args=(fecha_ini_str, fecha_fin_str), daemon=True)
        hilo_busqueda.start()

    def _tarea_buscar_glosas(self, fecha_ini, fecha_fin):
        try:
            if not self._iniciar_navegador_glosas_si_no_existe():
                return
            
            self.after(0, self._log_to_console, "[Glosas] Ejecutando búsqueda inicial...")
            
            resultados, estado_paginacion = glosas_downloader.fase_buscar(self.driver_glosas, fecha_ini, fecha_fin, self.base_path)
            
            self.after(0, self._log_to_console, f"✅ [Glosas] Búsqueda completada. {estado_paginacion['info_texto']}.")
            
            self.resultados_actuales = resultados
            self.after(0, self._actualizar_ui_resultados, resultados, estado_paginacion)

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Glosas] Error en la búsqueda: {e}")
            self.after(0, lambda: Messagebox.show_error(f"Ocurrió un error durante la búsqueda:\n\n{e}", "Error de Búsqueda"))
        finally:
            self.glosas_thread_lock.release()

    def _mostrar_resultados_glosas(self, resultados):
        for widget in self.results_frame.winfo_children():
            widget.destroy()

        if not resultados:
            Messagebox.show_info("No se encontraron glosas para el rango de fechas seleccionado.", "Búsqueda Vacía", parent=self)
            return
        
        self.results_frame.grid_rowconfigure(0, weight=1)
        self.results_frame.grid_columnconfigure(0, weight=1)

        cols = ["radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"]
        tree = ttk.Treeview(self.results_frame, columns=cols, show='headings', selectmode='extended')
        
        tree.heading("radicacion", text="N° Radicación")
        tree.heading("fecha_rad", text="Fecha Rad.")
        tree.heading("factura", text="N° Factura")
        tree.heading("valor_factura", text="Valor Factura")
        tree.heading("valor_glosado", text="Valor Glosado")
        tree.column("radicacion", width=200)
        tree.column("fecha_rad", width=100, anchor='center')

        scrollbar = ttk.Scrollbar(self.results_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        for res in resultados:
            tree.insert("", "end", iid=res['id'], values=(res['radicacion'], res['fecha_rad'], res['factura'], res['valor_factura'], res['valor_glosado']))

        tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        download_buttons_frame = ttk.Frame(self.results_frame)
        download_buttons_frame.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(10,0))
        
        def on_download(download_type):
            selected_iids = tree.selection()
            if not selected_iids:
                Messagebox.show_warning("No ha seleccionado ninguna glosa de la tabla.", "Selección Vacía", parent=self)
                return

            items_a_descargar = []
            for iid in selected_iids:
                item_data = tree.item(iid)
                id_cuenta = iid 
                numero_factura = item_data['values'][2]

                items_a_descargar.append({
                    "id": id_cuenta,
                    "factura": numero_factura,
                    "detalle": download_type == 'detalle',
                    "glosa": download_type == 'glosa'
                })
            
            tree.selection_set([]) 
            
            Messagebox.show_info(
                title="Descarga Iniciada",
                message=f"La descarga de {len(items_a_descargar)} item(s) ha comenzado en segundo plano.\nPuede seguir usando la aplicación.\n\nRevise el 'Registro de Proceso' en el panel de Configuración para ver el progreso.",
                parent=self
            )

            self._log_to_console(f"\n[Glosas] Se procederá a descargar '{download_type}' para {len(items_a_descargar)} item(s).")
            threading.Thread(target=self._tarea_descargar_glosas, args=(items_a_descargar,), daemon=True).start()

        btn_descargar_detalle = ttk.Button(download_buttons_frame, text="Descargar Detalles Seleccionados", bootstyle="info", command=lambda: on_download('detalle'))
        btn_descargar_detalle.pack(side=LEFT, expand=True, fill=X, padx=(0, 5))
        
        btn_descargar_glosa = ttk.Button(download_buttons_frame, text="Descargar Glosas Seleccionadas", bootstyle="info", command=lambda: on_download('glosa'))
        btn_descargar_glosa.pack(side=LEFT, expand=True, fill=X, padx=(5, 0))

    def _tarea_descargar_glosas(self, items_a_descargar):
        if not self.glosas_thread_lock.acquire(blocking=False):
            self.after(0, self._log_to_console, "⚠️ [Glosas] Ya hay una tarea en ejecución...")
            self.after(0, lambda: Messagebox.show_warning("Ya hay otra tarea en ejecución.", "Tarea en Progreso"))
            return

        try:
            if not self.driver_glosas:
                raise Exception("No se encontró un navegador activo para la descarga.")

            self.after(0, self._log_to_console, f"--- Iniciando lote de {len(items_a_descargar)} descargas ---")
            
            for item in items_a_descargar:
                processed_id = glosas_downloader.descargar_item_especifico(
                    self.driver_glosas, 
                    item,
                    self.download_dir_glosas,
                    self.last_processed_glosa_id
                )
                
                self.last_processed_glosa_id = processed_id

            self.after(0, self._log_to_console, "--- Lote de descargas finalizado. ---")
            self.after(0, lambda: Messagebox.show_info("La descarga de los archivos seleccionados ha finalizado.", "Descarga Completada"))

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Glosas] Error en la descarga: {e}")
            self.after(0, lambda: Messagebox.show_error(f"Ocurrió un error durante la descarga:\n\n{e}", "Error de Descarga"))
        finally:
            self.glosas_thread_lock.release()

    def _actualizar_ui_resultados(self, resultados, estado_paginacion):
        self._mostrar_resultados_glosas(resultados)

        if estado_paginacion:
            self.pagination_frame.grid()
            self.lbl_paginacion_info.config(text=estado_paginacion["info_texto"])
            
            self.btn_anterior.config(state="disabled" if estado_paginacion["anterior_deshabilitado"] else "normal")
            self.btn_siguiente.config(state="disabled" if estado_paginacion["siguiente_deshabilitado"] else "normal")

            valor_map = {"-1": "Todos"}
            valor_actual = estado_paginacion["entradas_actuales"]
            self.combo_entradas.set(valor_map.get(valor_actual, valor_actual))
        else:
            self.pagination_frame.grid_remove()

        self.update_idletasks()
        try:
            self.results_frame.tkraise()
            self.main_panel.focus_set()
        except:
            pass
        
    def on_navegar(self, direccion):
        if not self.glosas_thread_lock.acquire(blocking=False):
            Messagebox.show_warning("Ya hay una tarea en ejecución.", "Tarea en Progreso")
            return
            
        self._log_to_console(f"\n[Glosas] Solicitando página '{direccion}'...")
        threading.Thread(target=self._tarea_navegar, args=(direccion,), daemon=True).start()

    def _tarea_navegar(self, direccion):
        try:
            glosas_downloader.navegar_pagina(self.driver_glosas, direccion)
            resultados, estado_paginacion = glosas_downloader.extraer_datos_tabla_actual(self.driver_glosas)
            
            self.resultados_actuales = resultados
            self.after(0, self._actualizar_ui_resultados, resultados, estado_paginacion)
        except Exception as e:
            self.after(0, self._log_to_console, f"❌ Error al navegar: {e}")
        finally:
            self.glosas_thread_lock.release()

    def on_cambiar_entradas(self, event=None):
        if not self.glosas_thread_lock.acquire(blocking=False):
            Messagebox.show_warning("Ya hay una tarea en ejecución.", "Tarea en Progreso")
            return
        
        valor_map = {"Todos": "-1"}
        valor_seleccionado = self.combo_entradas.get()
        valor_real = valor_map.get(valor_seleccionado, valor_seleccionado)
        
        self._log_to_console(f"\n[Glosas] Solicitando mostrar '{valor_seleccionado}' entradas...")
        threading.Thread(target=self._tarea_cambiar_entradas, args=(valor_real,), daemon=True).start()

    def _tarea_cambiar_entradas(self, valor):
        try:
            glosas_downloader.cambiar_numero_entradas(self.driver_glosas, valor)
            resultados, estado_paginacion = glosas_downloader.extraer_datos_tabla_actual(self.driver_glosas)
            
            self.resultados_actuales = resultados
            self.after(0, self._actualizar_ui_resultados, resultados, estado_paginacion)
        except Exception as e:
            self.after(0, self._log_to_console, f"❌ Error al cambiar entradas: {e}")
        finally:
            self.glosas_thread_lock.release()

    def _consolidar_archivos_reporte(self, ruta_carpeta_reporte, fecha_ini, fecha_fin):
            """
            Encuentra todos los archivos Excel en la carpeta de reporte, los une
            en un único archivo 'Consolidado de Glosas.xlsx' con un encabezado personalizado.
            """
            try:
                self.after(0, self._log_to_console, f"[Consolidador] Buscando archivos Excel en '{os.path.basename(ruta_carpeta_reporte)}'...")
                
                # Buscar todos los archivos .xls y .xlsx en la carpeta
                archivos_excel = glob.glob(os.path.join(ruta_carpeta_reporte, "*.xls*"))

                if not archivos_excel:
                    self.after(0, self._log_to_console, "[Consolidador] No se encontraron archivos Excel para consolidar.")
                    return None

                self.after(0, self._log_to_console, f"[Consolidador] Se encontraron {len(archivos_excel)} archivos para procesar.")
                
                lista_de_dataframes = []
                for archivo in archivos_excel:
                    try:
                        # Leemos el excel, OMITIENDO LA PRIMERA FILA que es el título.
                        df_individual = pd.read_excel(archivo, skiprows=1)
                        if not df_individual.empty:
                            lista_de_dataframes.append(df_individual)
                    except Exception as e:
                        self.after(0, self._log_to_console, f"⚠️  [Consolidador] No se pudo leer el archivo '{os.path.basename(archivo)}': {e}")

                if not lista_de_dataframes:
                    self.after(0, self._log_to_console, "❌ [Consolidador] Ningún archivo contenía datos válidos para consolidar.")
                    return None
                
                # Unir todos los dataframes en uno solo
                df_consolidado = pd.concat(lista_de_dataframes, ignore_index=True)
                self.after(0, self._log_to_console, f"[Consolidador] Se consolidaron un total de {len(df_consolidado)} filas.")

                # --- Creación del archivo Excel con encabezado personalizado ---
                
                # 1. Preparar el contenido del encabezado personalizado
                encabezado_personalizado = pd.DataFrame({
                    'A': [f"Reporte de Glosas Recopilado"],
                    'B': [f"Período: {fecha_ini} a {fecha_fin}"],
                    'C': [f"Total Archivos Procesados: {len(archivos_excel)}"]
                })

                # 2. Definir el nombre y la ruta del archivo final
                nombre_archivo_final = "Consolidado de Glosas.xlsx"
                ruta_archivo_final = os.path.join(ruta_carpeta_reporte, nombre_archivo_final)

                # 3. Usar ExcelWriter para escribir en el archivo por partes
                with pd.ExcelWriter(ruta_archivo_final, engine='openpyxl') as writer:
                    # Escribir el encabezado personalizado sin su propio header
                    encabezado_personalizado.to_excel(writer, sheet_name='Consolidado', startrow=0, index=False, header=False)
                    
                    # Escribir el dataframe consolidado, dejando espacio para el encabezado
                    # startrow=3 deja una fila en blanco para mejor legibilidad
                    df_consolidado.to_excel(writer, sheet_name='Consolidado', startrow=3, index=False)
                
                self.after(0, self._log_to_console, f"✅ [Consolidador] Archivo final '{nombre_archivo_final}' creado exitosamente.")
                return nombre_archivo_final

            except Exception as e:
                self.after(0, self._log_to_console, f"❌ [Consolidador] Error fatal durante la consolidación: {e}")
                return None

    def _organizar_archivos_reporte(self, files_antes_de_descarga):
        try:
            self.after(0, self._log_to_console, "[Organizador] Identificando archivos nuevos...")

            files_despues_de_descarga = set(os.listdir(self.download_dir_glosas))
            archivos_nuevos = files_despues_de_descarga - files_antes_de_descarga

            # Ignorar temporales (Chrome/Edge)
            archivos_nuevos = {f for f in archivos_nuevos if not f.endswith(".crdownload") and not f.endswith(".tmp")}

            if not archivos_nuevos:
                self.after(0, self._log_to_console, "[Organizador] No se encontraron nuevos archivos para organizar.")
                return None, None

            fecha_actual_str = datetime.now().strftime("%Y-%m-%d")
            nombre_carpeta_reporte = f"Reporte de Glosas {fecha_actual_str}"
            ruta_carpeta_reporte = os.path.join(self.download_dir_glosas, nombre_carpeta_reporte)

            os.makedirs(ruta_carpeta_reporte, exist_ok=True)
            self.after(0, self._log_to_console, f"[Organizador] Carpeta de reporte creada: '{nombre_carpeta_reporte}'")

            for nombre_archivo in archivos_nuevos:
                ruta_origen = os.path.join(self.download_dir_glosas, nombre_archivo)
                if not os.path.isfile(ruta_origen):
                    continue
                ruta_destino = os.path.join(ruta_carpeta_reporte, nombre_archivo)

                # Evitar colisiones: si existe, versionar
                base, ext = os.path.splitext(ruta_destino)
                idx = 1
                while os.path.exists(ruta_destino):
                    ruta_destino = f"{base} ({idx}){ext}"
                    idx += 1

                shutil.move(ruta_origen, ruta_destino)

            self.after(0, self._log_to_console, f"[Organizador] Se movieron {len(archivos_nuevos)} archivos a la nueva carpeta.")
            return ruta_carpeta_reporte, nombre_carpeta_reporte

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Organizador] Error al mover archivos: {e}")
            return None, None


    # ########################################################################## #
    # ## INICIO DE LA NUEVA LÓGICA DE AUTOMATIZACIÓN (VERSIÓN SIMPLIFICADA)     ##
    # ########################################################################## #

    def iniciar_proceso_automatizacion_integrada(self):
        """
        Se activa con el botón 'Iniciar Automatización'.
        Pregunta al usuario cuántos de los ITEMS VISIBLES en la tabla desea descargar en cadena.
        """
        # 1. Bloquear para evitar acciones duplicadas
        if not self.glosas_thread_lock.acquire(blocking=False):
            Messagebox.show_warning("Ya hay una tarea de descarga en ejecución.", "Tarea en Progreso")
            return

        # 2. Verificar que tengamos resultados en la tabla para procesar
        if not self.resultados_actuales:
            Messagebox.show_warning("No hay resultados en la tabla para automatizar.", "Tabla Vacía", parent=self)
            self.glosas_thread_lock.release() # Liberamos el bloqueo si no hay nada que hacer
            return

        total_en_tabla = len(self.resultados_actuales)
        items_a_procesar = []

        try:
            # 3. Preguntar al usuario la cantidad
            cantidad_str = simpledialog.askstring(
                "Automatizar Descarga Secuencial",
                f"La tabla actual muestra {total_en_tabla} glosas.\n\n"
                f"¿Cuántas desea descargar en secuencia?\n"
                f"Escriba un número (1 a {total_en_tabla}) o la palabra 'todas'.",
                parent=self
            )

            # 4. Si el usuario cancela, liberamos el bloqueo y terminamos
            if not cantidad_str:
                self._log_to_console("[Automatización] Proceso cancelado por el usuario.")
                self.glosas_thread_lock.release()
                return

            # 5. Procesar la respuesta del usuario
            cantidad_str = cantidad_str.strip().lower()
            if cantidad_str == "todas":
                cantidad_a_descargar = total_en_tabla
            else:
                try:
                    cantidad_a_descargar = int(cantidad_str)
                    if not (0 < cantidad_a_descargar <= total_en_tabla):
                        Messagebox.show_error("Cantidad Inválida", f"Por favor, ingrese un número entre 1 y {total_en_tabla}.", parent=self)
                        self.glosas_thread_lock.release()
                        return
                except ValueError:
                    Messagebox.show_error("Entrada Inválida", "Por favor, ingrese un número válido o la palabra 'todas'.", parent=self)
                    self.glosas_thread_lock.release()
                    return

            # 6. Preparar la lista de ítems a descargar
            # Tomamos los primeros N resultados de la lista que ya tenemos en memoria
            resultados_para_descargar = self.resultados_actuales[:cantidad_a_descargar]
            
            for item_data in resultados_para_descargar:
                items_a_procesar.append({
                    "id": item_data['id'],
                    "factura": item_data['factura'],
                    "detalle": False,  # No descargamos detalle
                    "glosa": True      # Solo descargamos la glosa
                })

            # 7. Si todo es correcto, lanzamos la tarea en segundo plano
            # IMPORTANTE: No liberamos el bloqueo aquí. La tarea de fondo lo hará cuando termine.
            self._log_to_console(f"\n🤖 [Automatización] Iniciando descarga en cadena de {len(items_a_procesar)} glosas...")
            threading.Thread(
                target=self._tarea_ejecutar_descarga_en_cadena,
                args=(items_a_procesar,),
                daemon=True
            ).start()

        except Exception as e:
            # Si ocurre cualquier error inesperado durante la preparación, lo notificamos y liberamos el bloqueo
            self._log_to_console(f"❌ Error al preparar la automatización: {e}")
            self.glosas_thread_lock.release()

    def _tarea_ejecutar_descarga_en_cadena(self, items_a_descargar):
        """
        Descarga en cadena, organiza los archivos nuevos en una carpeta de reporte
        y consolida TODOS los campos disponibles de cada Excel en un solo archivo
        con formato estético. Luego genera el Informe Ejecutivo mapeando a
        (radicacion, fecha_rad, factura, valor_factura, valor_glosado).
        """
        try:
            if not self.driver_glosas:
                raise Exception("El navegador de automatización no está iniciado.")
            
            files_antes = set(os.listdir(self.download_dir_glosas))
            total = len(items_a_descargar)
            self.after(0, self._log_to_console, f"--- Procesando {total} glosas una por una ---")

            # Descarga en cadena
            for i, item in enumerate(items_a_descargar):
                self.after(0, self._log_to_console, f"[{i+1}/{total}] Procesando factura {item['factura']}...")
                processed_id = glosas_downloader.descargar_item_especifico(
                    self.driver_glosas, item, self.download_dir_glosas, self.last_processed_glosa_id
                )
                self.last_processed_glosa_id = processed_id

            self.after(0, self._log_to_console, "✅ --- Descarga en cadena finalizada. ---")

            # Organizar archivos nuevos
            ruta_carpeta_reporte, nombre_carpeta_reporte = self._organizar_archivos_reporte(files_antes)
            if not ruta_carpeta_reporte:
                mensaje_final = f"Se procesaron {total} glosas, pero no se encontraron nuevos archivos para organizar."
                self.after(0, lambda: Messagebox.show_info("Proceso Finalizado", mensaje_final, parent=self))
                return

            # =================== CONSOLIDACIÓN COMPLETA (todas las columnas) ===================
            # Mapeo de metadatos por factura tomado de self.resultados_actuales
            meta_por_factura = self._extraer_meta_por_factura(items_a_descargar)

            pattern = os.path.join(ruta_carpeta_reporte, "*.xls*")
            archivos_excel = glob.glob(pattern)

            if not archivos_excel:
                self.after(0, self._log_to_console, "❌ [Consolidación] No se encontraron Excel en la carpeta de reporte.")
                mensaje_final = f"📁 Carpeta creada: '{nombre_carpeta_reporte}', pero sin archivos Excel válidos."
                self.after(0, lambda: Messagebox.show_info("Proceso Finalizado", mensaje_final, parent=self))
                return

            dataframes = []
            for ruta in archivos_excel:
                try:
                    df, factura_detectada = self._leer_glosa_excel(ruta)
                    # Enriquecer con metadatos de resultados (si existen)
                    meta = meta_por_factura.get(str(factura_detectada), {})
                    df["factura"] = str(factura_detectada)
                    df["radicacion"] = meta.get("radicacion")
                    df["fecha_rad"] = pd.to_datetime(meta.get("fecha_rad"), errors="coerce") if meta.get("fecha_rad") else pd.NaT
                    df["valor_factura"] = meta.get("valor_factura")
                    # Si ya existe una columna de detalle "Valor Glosado", déjala, y además crea columna normalizada
                    # valor_glosado_item (numérica). El valor_glosado total lo mapeamos luego.
                    if "Valor Glosado" in df.columns:
                        # normalizar valor_glosado_item
                        def _to_float(x):
                            import re
                            import pandas as pd
                            if pd.isna(x): return 0.0
                            if isinstance(x, (int, float)): return float(x)
                            x = str(x)
                            x = re.sub(r"[^\d\-,.]", "", x)
                            x = x.replace(".", "").replace(",", ".")
                            try:
                                return float(x)
                            except:
                                return float("nan")
                        df["valor_glosado_item"] = df["Valor Glosado"].map(_to_float)
                    else:
                        df["valor_glosado_item"] = float("nan")

                    # valor_glosado total desde meta (si lo hay)
                    df["valor_glosado"] = meta.get("valor_glosado", None)

                    # Fechas del archivo (si trae) → formato fecha
                    if "Fecha" in df.columns:
                        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")

                    dataframes.append(df)
                except Exception as e:
                    self.after(0, self._log_to_console, f"⚠️ [Consolidación] {os.path.basename(ruta)}: {e}")

            if not dataframes:
                mensaje_final = (
                    f"Se procesaron {total} glosas.\n\n"
                    f"📁 Carpeta creada:\n'{nombre_carpeta_reporte}'\n\n"
                    f"Pero no se pudieron consolidar datos válidos."
                )
                self.after(0, lambda: Messagebox.show_info("Proceso Finalizado", mensaje_final, parent=self))
                return

            # Unión de columnas (todas las columnas presentes en algún archivo)
            # Orden: primero columnas 'meta' y luego el resto en orden legible
            columnas_meta = ["radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"]
            # Candidatas “bonitas” si existen
            preferidas = [
                "Id Glosa", "Id Item", "Descripcion Item", "Tipo", "Descripcion",
                "Justificacion", "Valor Glosado", "valor_glosado_item", "Usuario", "Fecha", "Estado", "Opciones"
            ]

            union_cols = set(columnas_meta)
            for df in dataframes:
                union_cols.update(df.columns.tolist())
            # Orden final: meta, luego preferidas que existan, luego el resto alfabético
            resto = [c for c in sorted(union_cols) if c not in columnas_meta and c not in preferidas]
            columnas_finales = [c for c in columnas_meta if c in union_cols] + \
                            [c for c in preferidas if c in union_cols] + \
                            resto

            # Reindex y concatenar
            dataframes = [df.reindex(columns=columnas_finales) for df in dataframes]
            df_consolidado_todo = pd.concat(dataframes, ignore_index=True)

            # Guardar consolidado estético
            nombre_consolidado = f"CONSOLIDADO_DETALLE_GLOSAS_{self.fecha_ini_actual}_a_{self.fecha_fin_actual}.xlsx"
            ruta_consolidado = os.path.join(ruta_carpeta_reporte, nombre_consolidado)
            try:
                self._guardar_consolidado_estetico(df_consolidado_todo, ruta_consolidado, hoja="Consolidado")
                self.after(0, self._log_to_console, f"📄 Consolidado guardado: {os.path.basename(ruta_consolidado)}")
            except Exception as e:
                self.after(0, self._log_to_console, f"⚠️ [Consolidación] No se pudo escribir el consolidado estético: {e}")
                ruta_consolidado = None

            # =================== PREPARAR DF PARA EL INFORME ===================
            # El informe necesita: radicacion, fecha_rad, factura, valor_factura, valor_glosado
            # Si no vino valor_glosado total en metadatos, calcularlo sumando los 'valor_glosado_item' por factura.
            df_info = df_consolidado_todo.copy()

            # Asegurar columna valor_glosado (total por factura)
            if df_info.get("valor_glosado") is None or df_info["valor_glosado"].isna().all():
                # Agrupar por factura usando la suma de valor_glosado_item
                suma_glosa = (df_info.groupby("factura", as_index=False)["valor_glosado_item"]
                            .sum()
                            .rename(columns={"valor_glosado_item": "valor_glosado"}))
                # Tomar una fila por factura para traer radicacion/fecha_rad/valor_factura
                base_fact = (df_info[["factura", "radicacion", "fecha_rad", "valor_factura"]]
                            .drop_duplicates(subset=["factura"], keep="first"))
                df_para_informe = base_fact.merge(suma_glosa, on="factura", how="left")
            else:
                # Ya viene valor_glosado desde metadatos: tomar primera fila por factura
                df_para_informe = (df_info[["radicacion", "fecha_rad", "factura", "valor_factura", "valor_glosado"]]
                                .drop_duplicates(subset=["factura"], keep="first"))

            # Generar Informe Ejecutivo
            nombre_informe = f"INFORME_GLOSAS_{self.fecha_ini_actual}_a_{self.fecha_fin_actual}.xlsx"
            ruta_informe = os.path.join(ruta_carpeta_reporte, nombre_informe)

            try:
                informe_path = self._crear_informe_ejecutivo(
                    df_para_informe, self.fecha_ini_actual, self.fecha_fin_actual, ruta_informe, self.logos
                )
                mensaje_final = (
                    f"Se han procesado y descargado {total} glosas exitosamente.\n\n"
                    f"📁 Carpeta: '{nombre_carpeta_reporte}'\n\n"
                    + (f"📄 Consolidado generado: '{os.path.basename(ruta_consolidado)}'\n" if ruta_consolidado else "")
                    + f"📊 Informe Ejecutivo generado: '{os.path.basename(informe_path)}'"
                )
            except Exception as e:
                self.after(0, self._log_to_console, f"❌ [Informe] Error generando informe: {e}")
                mensaje_final = (
                    f"Se procesaron {total} glosas y se organizó la carpeta:\n'{nombre_carpeta_reporte}'.\n\n"
                    f"Pero ocurrió un error al generar el informe."
                )

            self.after(0, lambda: Messagebox.show_info("Proceso Finalizado", mensaje_final, parent=self))

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Automatización] Error durante la descarga en cadena: {e}")
            self.after(0, lambda: Messagebox.show_error("Error en Automatización", f"Ocurrió un error durante el proceso:\n\n{e}", parent=self))
        finally:
            self.glosas_thread_lock.release()
            self._log_to_console("[Automatización] Proceso terminado y bloqueo liberado.")



    # ########################################################################## #
    # ## FIN DE LA NUEVA LÓGICA DE AUTOMATIZACIÓN (VERSIÓN SIMPLIFICADA)        ##
    # ########################################################################## #


    def _extraer_meta_por_factura(self, items_a_descargar):
        """
        Construye un diccionario {factura: meta} con campos:
        radicacion, fecha_rad, valor_factura, valor_glosado.
        Se apoya en self.resultados_actuales (que viene de la tabla de resultados).
        """
        meta = {}
        # indexar resultados por factura
        idx = {str(r.get("factura")): r for r in (self.resultados_actuales or [])}
        for it in items_a_descargar:
            fac = str(it.get("factura"))
            r = idx.get(fac, {})
            meta[fac] = {
                "radicacion": r.get("radicacion"),
                "fecha_rad": r.get("fecha_rad"),
                "valor_factura": r.get("valor_factura"),
                "valor_glosado": r.get("valor_glosado"),
            }
        return meta


    def _leer_glosa_excel(self, ruta_archivo):
        """
        Lee un archivo 'Glosas de la cuenta ####.xls[x]' detectando la fila de cabecera.
        Devuelve (df, factura_detectada).
        - Si la primera celda de la fila 0 empieza por 'Glosas de la cuenta', salta esa fila.
        - Normaliza algunos nombres comunes (no altera los originales para el consolidado total).
        """
        import re
        ext = os.path.splitext(ruta_archivo)[1].lower()
        # Para .xls se requiere xlrd<=1.2.0
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"

        # Detectar factura por nombre de archivo
        base = os.path.basename(ruta_archivo)
        m = re.search(r"(\d+)", base)
        factura_detectada = m.group(1) if m else ""

        # Leer una muestra para detectar título
        df_head = pd.read_excel(ruta_archivo, engine=engine, header=None, nrows=2)
        primera = str(df_head.iat[0,0]).strip().lower() if df_head.shape[1] > 0 else ""
        skip = 1 if primera.startswith("glosas de la cuenta") else 0

        # Leer definitivo (si hay título, skip=1 → usa la fila 2 como encabezado)
        df = pd.read_excel(ruta_archivo, engine=engine, skiprows=skip)

        # Limpiar nombres de columnas (quitar espacios extras)
        df.columns = [str(c).strip() for c in df.columns]

        return df, factura_detectada


    def _guardar_consolidado_estetico(self, df: pd.DataFrame, ruta_salida: str, hoja: str = "Consolidado"):
        """
        Guarda un DataFrame en Excel con formato:
        - Encabezado en negrita + fondo gris
        - Congela fila de encabezado
        - Autofiltro
        - Autoajuste de anchos
        - Formato de moneda para columnas relevantes
        - Formato fecha en columnas de fecha
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = hoja

        # Escribir DF
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)
            # Encabezado
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")

        # Congelar encabezado y autofiltro
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Formatos de moneda
        cols_monedas = {"valor_factura", "valor_glosado", "valor_glosado_item"}
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            header = ws.cell(row=1, column=col[0].column).value
            if header and str(header).strip() in cols_monedas:
                for c in col:
                    c.number_format = '#,##0'

        # Formatos de fecha
        cols_fecha = {"fecha_rad", "Fecha"}
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            header = ws.cell(row=1, column=col[0].column).value
            if header and str(header).strip() in cols_fecha:
                for c in col:
                    # Si es datetime, Excel lo mostrará con formato corto
                    c.number_format = 'yyyy-mm-dd hh:mm' if str(header).strip() == "Fecha" else 'yyyy-mm-dd'

        # Autoajuste simple de anchos
        for column_cells in ws.columns:
            length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    val = "" if cell.value is None else str(cell.value)
                    length = max(length, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(length + 2, 60)

        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        wb.save(ruta_salida)



    def _reconfigurar_directorio_descarga(self, nuevo_directorio):
        """Reconfigura el directorio de descarga del navegador existente."""
        try:
            self.driver_glosas.execute_cdp_cmd('Page.setDownloadBehavior', {
                'behavior': 'allow',
                'downloadPath': nuevo_directorio
            })
            self.after(0, self._log_to_console, f"[Config] Directorio de descarga cambiado a: {os.path.basename(nuevo_directorio)}")
        except Exception as e:
            self.after(0, self._log_to_console, f"[Config] Advertencia: No se pudo cambiar directorio de descarga: {e}")

    def _consolidar_archivos_excel_desde_carpeta(self, ruta_carpeta, fecha_ini, fecha_fin):
        """Consolida todos los .xls/.xlsx de la carpeta y devuelve la RUTA del archivo creado."""
        try:
            pattern = os.path.join(ruta_carpeta, "*.xls*")
            archivos_excel = glob.glob(pattern)

            self.after(0, self._log_to_console, f"[Consolidación] Archivos encontrados: {len(archivos_excel)}")
            if not archivos_excel:
                raise Exception("No se encontraron archivos Excel para consolidar.")

            datos_consolidados = []
            for archivo in archivos_excel:
                try:
                    df = self._leer_excel_robusto(archivo)
                    # Intento normalizar; si falla por encabezado en fila 2, reintento con skiprows=1
                    try:
                        df_norm = self._normalizar_columnas(df)
                    except RuntimeError as e:
                        if "Faltan columnas requeridas" in str(e):
                            ext = os.path.splitext(archivo)[1].lower()
                            if ext == ".xlsx":
                                df2 = pd.read_excel(archivo, engine="openpyxl", skiprows=1)
                            elif ext == ".xls":
                                try:
                                    import xlrd  # xlrd<=1.2.0
                                    df2 = pd.read_excel(archivo, engine="xlrd", skiprows=1)
                                except Exception as e2:
                                    raise RuntimeError(f"No se pudo reintentar con skiprows=1 en {os.path.basename(archivo)}: {e2}")
                            else:
                                raise
                            df_norm = self._normalizar_columnas(df2)
                        else:
                            raise
                    if not df_norm.empty:
                        datos_consolidados.append(df_norm)
                except Exception as e:
                    self.after(0, self._log_to_console, f"⚠️ [Consolidación] {os.path.basename(archivo)}: {e}")

            if not datos_consolidados:
                raise Exception("Ningún archivo Excel pudo ser procesado.")

            df_consolidado = pd.concat(datos_consolidados, ignore_index=True)

            nombre_archivo = f"CONSOLIDADO_GLOSAS_{fecha_ini}_a_{fecha_fin}.xlsx"
            ruta_consolidado = os.path.join(ruta_carpeta, nombre_archivo)

            df_consolidado.to_excel(ruta_consolidado, index=False, engine="openpyxl")
            self.after(0, self._log_to_console, f"✅ [Consolidación] {os.path.basename(ruta_consolidado)} con {len(df_consolidado):,} filas.")
            return ruta_consolidado

        except Exception as e:
            self.after(0, self._log_to_console, f"❌ [Consolidación] {e}")
            return None



    # ########################################################################## #
    # ## FIN DE LA NUEVA LÓGICA DE AUTOMATIZACIÓN INTEGRADA                     ##
    # ########################################################################## #

    # ───────────────────────── UTILIDADES DE LECTURA/FORMATO ─────────────────────────

    def _leer_excel_robusto(self, ruta_archivo: str) -> pd.DataFrame:
        """
        Lee .xlsx con openpyxl y .xls con xlrd (si está disponible y soportado).
        Lanza excepción clara si no es posible.
        """
        ext = os.path.splitext(ruta_archivo)[1].lower()
        try:
            if ext == ".xlsx":
                return pd.read_excel(ruta_archivo, engine="openpyxl")
            elif ext == ".xls":
                try:
                    import xlrd  # xlrd<=1.2.0 requerido para .xls
                    return pd.read_excel(ruta_archivo, engine="xlrd")
                except Exception as e:
                    raise RuntimeError(
                        f"Para leer .xls necesitas xlrd<=1.2.0 o convertir a .xlsx. Archivo: {os.path.basename(ruta_archivo)} | Error: {e}"
                    )
            else:
                raise RuntimeError(f"Extensión no soportada: {ext} en {os.path.basename(ruta_archivo)}")
        except Exception as e:
            raise RuntimeError(f"No se pudo leer '{os.path.basename(ruta_archivo)}': {e}")

    def _norm_token(self, s: str) -> str:
        if s is None:
            return ""
        if not isinstance(s, str):
            s = str(s)
        # quita acentos y símbolos raros
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        s = s.lower()
        s = s.replace("\n", " ").replace("\r", " ")
        # normaliza variantes de numero: nro / n° / nº / no.
        s = re.sub(r"\b(nro|n°|nº|no)\b", " num ", s)
        # deja solo letras/numeros/espacio
        s = re.sub(r"[^\w\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _leer_excel_con_header_detectado(self, ruta_archivo: str) -> pd.DataFrame:
        """Lee cualquier hoja buscando la fila de encabezados (robusto)."""
        ext = os.path.splitext(ruta_archivo)[1].lower()
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        if ext == ".xls":
            try:
                import xlrd  # xlrd<=1.2.0 requerido
            except Exception as e:
                raise RuntimeError(f"Para .xls se requiere xlrd<=1.2.0. {e}")

        xl = pd.ExcelFile(ruta_archivo, engine=engine)
        ultimo_error = None

        for sh in xl.sheet_names:
            try:
                tmp = xl.parse(sh, header=None, dtype=str)
                # busca fila encabezado en las primeras 15 filas
                header_row = None
                lim = min(15, len(tmp))
                for i in range(lim):
                    fila = tmp.iloc[i].dropna().tolist()
                    linea = " ".join(self._norm_token(x) for x in fila)
                    score = 0
                    if re.search(r"radic|radicado|cuenta", linea): score += 1
                    if re.search(r"factur", linea): score += 1
                    if re.search(r"glos", linea): score += 1
                    if re.search(r"valor", linea): score += 1
                    if score >= 2:
                        header_row = i
                        break
                if header_row is None:
                    # cae por defecto al header=0
                    df = xl.parse(sh, header=0)
                else:
                    df = xl.parse(sh, header=header_row)
                # si la primera fila aún es título (1 celda no nula), descártala
                if df.shape[1] == 1 and df.shape[0] > 1:
                    continue
                return df
            except Exception as e:
                ultimo_error = e
                continue

        raise RuntimeError(f"No se pudo detectar encabezado en '{os.path.basename(ruta_archivo)}': {ultimo_error}")

    def _normalizar_columnas(self, df: pd.DataFrame, ruta_archivo: str = None) -> pd.DataFrame:
        """
        Mapea columnas variantes a un esquema estándar y tipifica.
        Soporta encabezados con acentos/símbolos y nombres alternos.
        Columnas estándar: radicacion, fecha_rad, factura, valor_factura, valor_glosado
        """
        # normaliza nombres
        col_norm = {c: self._norm_token(c) for c in df.columns}

        # heurística de mapeo
        colmap = {}
        for orig, norm in col_norm.items():
            if any(k in norm for k in ["radic", "radicado", "radicacion", "id_cuenta", "id cuenta", "cuenta"]):
                colmap[orig] = "radicacion"
            elif "fact" in norm and "num" in norm:
                colmap[orig] = "factura"
            elif norm == "factura" or ("factura" in norm and "prefijo" not in norm):
                colmap[orig] = "factura"
            elif "fecha" in norm and ("rad" in norm or "fac" in norm or True):
                colmap[orig] = "fecha_rad"
            elif "valor" in norm and "glos" in norm:
                colmap[orig] = "valor_glosado"
            elif ("vlr" in norm and "glos" in norm):
                colmap[orig] = "valor_glosado"
            elif "valor" in norm and "fact" in norm:
                colmap[orig] = "valor_factura"
            elif norm == "glosa":
                colmap[orig] = "valor_glosado"

        df2 = df.rename(columns=colmap)

        # completa ausentes con defaults
        # radicacion: desde nombre de archivo si no viene en la hoja
        if "radicacion" not in df2.columns:
            if ruta_archivo:
                m = re.search(r"(\d{5,})", os.path.basename(ruta_archivo))
                if m:
                    df2["radicacion"] = m.group(1)
        if "fecha_rad" not in df2.columns:
            df2["fecha_rad"] = pd.NA
        if "factura" not in df2.columns:
            # algunos reportes traen "num factura", "no factura" en otra forma
            # intenta inferir usando cualquier columna que contenga "fact" en el nombre original
            candidatos = [c for c, n in col_norm.items() if "fact" in n and c not in df2.columns]
            if candidatos:
                df2["factura"] = df[candidatos[0]].astype(str)
            else:
                df2["factura"] = pd.NA
        if "valor_factura" not in df2.columns:
            df2["valor_factura"] = 0
        if "valor_glosado" not in df2.columns:
            # si hay columna "glosa" o similar en números, úsala
            candidatos = [c for c, n in col_norm.items() if n == "glosa" or ("glos" in n and "valor" in n)]
            if candidatos:
                df2["valor_glosado"] = df[candidatos[0]]
            else:
                df2["valor_glosado"] = 0

        # conserva solo las estándar (en ese orden)
        for col in ["radicacion","fecha_rad","factura","valor_factura","valor_glosado"]:
            if col not in df2.columns:
                df2[col] = pd.NA
        df2 = df2[["radicacion","fecha_rad","factura","valor_factura","valor_glosado"]].copy()

        # tipificación
        df2["radicacion"] = df2["radicacion"].astype(str).str.strip()
        df2["factura"]    = df2["factura"].astype(str).str.strip()

        # fecha
        df2["fecha_rad"] = pd.to_datetime(df2["fecha_rad"], errors="coerce")

        # moneda → float
        def _to_float(x):
            if pd.isna(x): return 0.0
            if isinstance(x, (int, float)): return float(x)
            x = str(x)
            x = re.sub(r"[^\d\-,.]", "", x)               # quita símbolos
            x = x.replace(".", "").replace(",", ".")      # 1.234.567,89 → 1234567.89
            try:
                return float(x)
            except:
                return float("nan")

        df2["valor_factura"] = df2["valor_factura"].map(_to_float).fillna(0.0)
        df2["valor_glosado"] = df2["valor_glosado"].map(_to_float).fillna(0.0)

        return df2



    def _evaluar_calidad_datos(self, df: pd.DataFrame) -> dict:
        """
        Devuelve incidencias: duplicados, nulos, fechas inválidas, importes negativos.
        """
        incidencias = {}

        # Duplicados por (radicacion, factura)
        dup_mask = df.duplicated(subset=["radicacion","factura"], keep=False)
        incidencias["duplicados"] = df[dup_mask].sort_values(["radicacion","factura"]) if dup_mask.any() else pd.DataFrame()

        # Nulos
        nulos = {}
        for c in ["radicacion","fecha_rad","factura","valor_factura","valor_glosado"]:
            nulos[c] = int(df[c].isna().sum())
        incidencias["nulos"] = nulos

        # Fechas inválidas
        incidencias["fechas_invalidas"] = df[df["fecha_rad"].isna()]

        # Importes negativos
        incidencias["negativos"] = df[(df["valor_factura"] < 0) | (df["valor_glosado"] < 0)]

        return incidencias


    def _kpis_glosas(self, df: pd.DataFrame) -> dict:
        total_fact = float(df["valor_factura"].sum())
        total_glosa = float(df["valor_glosado"].sum())
        pct_glosa = (total_glosa / total_fact * 100.0) if total_fact > 0 else 0.0
        kpis = {
            "total_facturado": total_fact,
            "total_glosado": total_glosa,
            "porcentaje_glosa": pct_glosa,
            "n_facturas": df["factura"].nunique(),
            "n_radicaciones": df["radicacion"].nunique(),
        }
        return kpis


    def _tablas_resumen(self, df: pd.DataFrame) -> dict:
        # Top 10 facturas por valor glosado
        top_facturas = (df.groupby("factura", as_index=False)["valor_glosado"]
                        .sum()
                        .sort_values("valor_glosado", ascending=False)
                        .head(10))

        # Serie diaria
        diario = (df.assign(fecha=df["fecha_rad"].dt.date)
                    .groupby("fecha", as_index=False)[["valor_factura","valor_glosado"]]
                    .sum()
                    .sort_values("fecha"))

        return {"top_facturas": top_facturas, "diario": diario}

    def _formato_moneda_cop(self):
        # Formato COP: muestra separadores y símbolo opcional (puedes ajustar a "[$COP] #,##0")
        return '#,##0'

    def _autoajustar_columnas(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    def _agregar_tabla_df(self, ws, df: pd.DataFrame, start_row=1, start_col=1, header_fill="FFDDDDDD"):
        # Escribe DataFrame respetando encabezados
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Encabezado con estilo
                if r_idx == start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor=header_fill)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")

    def _crear_informe_ejecutivo(self, df_detalle: pd.DataFrame, fecha_ini: str, fecha_fin: str, ruta_salida: str, logos: dict = None) -> str:
        """
        Genera archivo Excel profesional con:
        - Portada
        - Resumen Ejecutivo (KPIs, gráficos)
        - Detalle (datos)
        - Calidad de Datos
        """
        wb = Workbook()

        # ───────────── Portada ─────────────
        ws_cover = wb.active
        ws_cover.title = "Portada"
        titulo = "INFORME EJECUTIVO DE GLOSAS"
        sub = f"Período: {fecha_ini} a {fecha_fin}"
        gen = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        autor = f"Usuario: {self.current_user.get('nombre','N/A')} - {self.current_user.get('cargo','N/A')}"

        ws_cover.merge_cells("A2:G2")
        ws_cover["A2"] = titulo
        ws_cover["A2"].font = Font(size=20, bold=True, color="1F497D")
        ws_cover["A2"].alignment = Alignment(horizontal="center")

        ws_cover.merge_cells("A4:G4")
        ws_cover["A4"] = sub
        ws_cover["A4"].font = Font(size=12, italic=True)
        ws_cover["A4"].alignment = Alignment(horizontal="center")

        ws_cover["A6"] = gen
        ws_cover["A7"] = autor

        # (Opcional) Insertar logos si los tienes en self.image_path
        # from openpyxl.drawing.image import Image as XLImage
        # try:
        #     if logos and logos.get("huv"):
        #         img = XLImage(os.path.join(self.image_path, "logo1.jpg"))
        #         img.width, img.height = 160, 160
        #         ws_cover.add_image(img, "F2")
        # except Exception:
        #     pass

        # ───────────── Resumen ─────────────
        ws_sum = wb.create_sheet("Resumen Ejecutivo")
        kpis = self._kpis_glosas(df_detalle)
        tablas = self._tablas_resumen(df_detalle)

        # KPIs
        kpi_items = [
            ("Total Facturado (COP)", kpis["total_facturado"]),
            ("Total Glosado (COP)", kpis["total_glosado"]),
            ("% Glosa", kpis["porcentaje_glosa"]),
            ("# Facturas", kpis["n_facturas"]),
            ("# Radicaciones", kpis["n_radicaciones"]),
        ]
        ws_sum["A1"] = "KPIs"
        ws_sum["A1"].font = Font(size=14, bold=True)
        row = 3
        for label, val in kpi_items:
            ws_sum[f"A{row}"] = label
            ws_sum[f"A{row}"].font = Font(bold=True)
            ws_sum[f"B{row}"] = float(val)
            if "COP" in label:
                ws_sum[f"B{row}"].number_format = self._formato_moneda_cop()
            elif "%" in label:
                ws_sum[f"B{row}"].number_format = "0.00%"
                ws_sum[f"B{row}"].value = float(val) / 100.0
            row += 1

        # Top 10 facturas (tabla + barra)
        ws_sum["D1"] = "Top 10 facturas por valor glosado"
        ws_sum["D1"].font = Font(size=14, bold=True)
        self._agregar_tabla_df(ws_sum, tablas["top_facturas"], start_row=3, start_col=4)
        # Formato moneda a la columna "valor_glosado"
        col_glosa = 4 + list(tablas["top_facturas"].columns).index("valor_glosado")  # columna absoluta
        for r in range(4, 4 + len(tablas["top_facturas"]) + 1):
            ws_sum.cell(row=r, column=col_glosa).number_format = self._formato_moneda_cop()

        # Gráfico barras
        chart = BarChart()
        chart.title = "Glosa por Factura (Top 10)"
        data_ref = Reference(ws_sum, min_col=col_glosa, min_row=3, max_row=3 + len(tablas["top_facturas"]))
        cats_ref = Reference(ws_sum, min_col=4, min_row=4, max_row=3 + len(tablas["top_facturas"]))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 12
        chart.width = 28
        ws_sum.add_chart(chart, "D16")

        # Serie diaria (tabla + línea)
        ws_sum["A10"] = "Evolución Diaria"
        ws_sum["A10"].font = Font(size=14, bold=True)
        diario = tablas["diario"].copy()
        # Asegurar tipos simples
        diario["fecha"] = diario["fecha"].astype(str)
        self._agregar_tabla_df(ws_sum, diario, start_row=12, start_col=1)
        # Formato moneda
        col_fact = 1 + list(diario.columns).index("valor_factura")
        col_glo = 1 + list(diario.columns).index("valor_glosado")
        for r in range(13, 13 + len(diario)):
            ws_sum.cell(row=r, column=col_fact).number_format = self._formato_moneda_cop()
            ws_sum.cell(row=r, column=col_glo).number_format = self._formato_moneda_cop()

        # Gráfico línea
        line = LineChart()
        line.title = "Tendencia diaria"
        data = Reference(ws_sum, min_col=col_fact, min_row=12, max_col=col_glo, max_row=12 + len(diario))
        cats = Reference(ws_sum, min_col=1, min_row=13, max_row=12 + len(diario))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 12
        line.width = 28
        ws_sum.add_chart(line, "A16")

        self._autoajustar_columnas(ws_sum)

        # ───────────── Detalle ─────────────
        ws_det = wb.create_sheet("Detalle")
        self._agregar_tabla_df(ws_det, df_detalle, start_row=1, start_col=1)
        ws_det.freeze_panes = "A2"
        # Formateo de columnas monetarias
        for r in range(2, 2 + len(df_detalle)):
            ws_det.cell(row=r, column= list(df_detalle.columns).index("valor_factura")+1 ).number_format = self._formato_moneda_cop()
            ws_det.cell(row=r, column= list(df_detalle.columns).index("valor_glosado")+1 ).number_format = self._formato_moneda_cop()
        self._autoajustar_columnas(ws_det)
        ws_det.auto_filter.ref = ws_det.dimensions

        # ───────────── Calidad de Datos ─────────────
        ws_q = wb.create_sheet("Calidad de Datos")
        ws_q["A1"] = "Incidencias de Datos"
        ws_q["A1"].font = Font(size=14, bold=True)

        incid = self._evaluar_calidad_datos(df_detalle)

        ws_q["A3"] = "Duplicados (radicacion, factura)"
        ws_q["A3"].font = Font(bold=True)
        dup = incid["duplicados"]
        if len(dup) > 0:
            self._agregar_tabla_df(ws_q, dup, start_row=4, start_col=1)
        else:
            ws_q["A4"] = "Sin duplicados detectados."

        start_r = 6 + len(dup) if len(dup) > 0 else 6
        ws_q[f"A{start_r}"] = "Nulos"
        ws_q[f"A{start_r}"].font = Font(bold=True)
        r = start_r + 1
        for k, v in incid["nulos"].items():
            ws_q[f"A{r}"] = k
            ws_q[f"B{r}"] = v
            r += 1

        r += 1
        ws_q[f"A{r}"] = "Fechas inválidas"
        ws_q[f"A{r}"].font = Font(bold=True)
        r += 1
        if len(incid["fechas_invalidas"]) > 0:
            self._agregar_tabla_df(ws_q, incid["fechas_invalidas"], start_row=r, start_col=1)
            r += len(incid["fechas_invalidas"]) + 3
        else:
            ws_q[f"A{r}"] = "Sin fechas inválidas."
            r += 3

        ws_q[f"A{r}"] = "Importes negativos"
        ws_q[f"A{r}"].font = Font(bold=True)
        r += 1
        if len(incid["negativos"]) > 0:
            self._agregar_tabla_df(ws_q, incid["negativos"], start_row=r, start_col=1)
        else:
            ws_q[f"A{r}"] = "Sin importes negativos."

        self._autoajustar_columnas(ws_q)

        # Guardar
        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        wb.save(ruta_salida)
        return ruta_salida



# --- PUNTO DE ENTRADA PRINCIPAL ---
if __name__ == "__main__":
    import multiprocessing
    import sys
    import argparse
    import tkinter as tk
    from tkinter import messagebox
    multiprocessing.freeze_support()
    parser = argparse.ArgumentParser(description="Gestor Coosalud para EVARISIS.")    
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument("--run-server", action="store_true", help="Ejecuta la lógica del servidor en segundo plano.")
    mode_group.add_argument("--run-client", action="store_true", help="Ejecuta la lógica del cliente en segundo plano.")
    parser.add_argument("--base-path", type=str, help="Ruta base necesaria para los modos de servidor/cliente.")
    parser.add_argument("--usuario", default="UsuarioDesconocido", type=str, help="Nombre del usuario para el servidor.")
    parser.add_argument("--lanzado-por-evarisis", action="store_true", help="Bandera de seguridad para el lanzamiento de la GUI.")
    parser.add_argument("--nombre", default="Usuario Invitado", type=str, help="Nombre del usuario.")
    parser.add_argument("--cargo", default="N/A", type=str, help="Cargo del usuario.")
    parser.add_argument("--foto", default="SIN_FOTO", type=str, help="Ruta a la foto del usuario.")
    parser.add_argument("--tema", default="litera", type=str, help="Tema de ttkbootstrap para la GUI.")
    parser.add_argument("--ruta-datos", type=str, help="Ruta a la carpeta de datos central (consistencia).")
    args = parser.parse_args()
    if args.run_server:
        # Lógica para el servidor (importar y ejecutar tray_app)
        if not args.base_path:
            print("ERROR: El modo --run-server requiere el argumento --base-path.")
            sys.exit(1)
        import tray_app
        tray_app.main(args.base_path, args.usuario)
        sys.exit(0)
    elif args.run_client:
        if not args.base_path:
            print("ERROR: El modo --run-client requiere el argumento --base-path.")
            sys.exit(1)
        import session_cliente
        session_cliente.run_client_logic(args.base_path)
        sys.exit(0)    
    else:
        if not args.lanzado_por_evarisis:
            root = tk.Tk()
            root.withdraw()
            messagebox.showwarning(
                "Acceso Denegado", 
                "Esta aplicación es un módulo de EVARISIS y debe ser ejecutada desde el dashboard principal."
            )
            root.destroy()
            sys.exit(1)
        app = CoosaludApp(
            nombre_usuario=args.nombre,
            cargo_usuario=args.cargo,
            foto_path=args.foto,
            tema=args.tema
        )
        app.mainloop()
